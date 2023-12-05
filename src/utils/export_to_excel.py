import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def auto_fit_columns(file_path):
    """
    Auto-fit column widths in an Excel file.
    """
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    for col in worksheet.iter_cols():  # type: ignore
        max_length = 0
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                print(f"Error processing cell {cell.coordinate}: {e}")

        adjusted_width = max_length + 2
        worksheet.column_dimensions[  # type: ignore
            get_column_letter(col[0].column)
        ].width = adjusted_width

    workbook.save(file_path)


def export_summary_statistics(statistics, output_file_path="summary_statistics.xlsx"):
    """
    Exports summary statistics to an Excel file.

    :param statistics: A list of dictionaries containing summary statistics for each generation.
    :param output_file_path: The path where the Excel file will be saved.
    """
    try:
        # Convert the statistics to a DataFrame for easy export to Excel
        df = pd.DataFrame(statistics)

        # Calculating distribution as a string for each generation
        df["distribution"] = df["distribution"].apply(
            lambda d: f"MWF: {d['MWF']}, TR: {d['TR']}"
        )

        # Reordering columns as per the new format
        df = df[
            [
                "generation",
                "total_courses",
                "distribution",
                "teacher_preference_adherence",
                "teacher_satisfaction",
                "average_fitness",
                "max_fitness",
                "preference_violations",
                "course_assignment_duplicates",
            ]
        ]

        # Writing to an Excel file
        df.to_excel(output_file_path, index=False)

        # auto_fit_columns widths
        auto_fit_columns(output_file_path)

        print(f"Summary statistics successfully exported to '{output_file_path}'.")

    except Exception as e:
        print(f"An error occurred during export: {e}")


def export_to_excel(
    best_chromosome, time_slots_details, output_file_path="data/final_schedule.xlsx"
):
    """
    Exports the best chromosome (schedule) to an Excel file.

    :param best_chromosome: The best chromosome to be exported.
    :param time_slots_details: Dictionary or similar structure containing the details of the time slots.
    :param output_file_path: The path where the Excel file will be saved.
    """
    try:
        data = []
        for gene in best_chromosome.genes:
            teacher_id = gene[3]
            course_id = gene[0]["Course Section ID"]
            time_slot_id = gene[2]["Time Slot ID"]
            room = gene[1]["Room Number"]

            # Lookup the full details of the time slot
            time_slot_detail = time_slots_details.get(time_slot_id, "Unknown Time Slot")

            data.append([teacher_id, course_id, time_slot_detail, room])

        columns = ["Teacher ID", "Course ID", "Time Slot", "Room"]
        df = pd.DataFrame(data, columns=columns)

        df.to_excel(output_file_path, index=False)

        # Auto fit columns
        auto_fit_columns(output_file_path)

        print(f"Final schedule successfully exported to '{output_file_path}'.")

    except Exception as e:
        print(f"An error occurred during export: {e}")
